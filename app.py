# app.py
# streamlit run app.py

import re
import io
import pandas as pd
import streamlit as st
import plotly.express as px
import folium
from streamlit_folium import st_folium

# Excel bonito (requiere openpyxl en requirements.txt)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =========================
# CONFIG
# =========================
st.set_page_config(page_title="CR | Mapas y Programas", page_icon="🛰️", layout="wide")

# =========================
# ESTILO (MISMA LÍNEA VISUAL, PERO SIN ROMPER TABS NI TEXTO)
# =========================
st.markdown(
    """
    <style>
      .block-container {padding-top: 1.0rem; padding-bottom: 2rem;}

      /* ✅ Evita transformaciones raras de texto (EstrukturAs, etc.) */
      h1, h2, h3, h4, h5, h6, p, span, div {
        text-transform: none !important;
        letter-spacing: normal !important;
        font-variant: normal !important;
      }

      .title{font-size: 28px; font-weight: 900; letter-spacing: -0.02em; margin-bottom: 2px;}
      .subtitle{color:#6b7280; margin-top:0px; margin-bottom: 14px;}
      .kpi-grid{display:flex; gap:14px; flex-wrap:wrap; margin-bottom: 10px;}

      /* === KPI con color === */
      .kpi{
        border-radius:18px;
        padding:16px 18px;
        box-shadow:0 10px 25px rgba(0,0,0,0.08);
        min-width: 240px; flex:1;
        border:1px solid rgba(255,255,255,0.14);
        color:#0b1220;
      }
      .kpi.kpi-a{
        background: linear-gradient(135deg, rgba(99,102,241,0.22), rgba(59,130,246,0.10));
        border:1px solid rgba(99,102,241,0.35);
      }
      .kpi.kpi-b{
        background: linear-gradient(135deg, rgba(16,185,129,0.22), rgba(34,197,94,0.10));
        border:1px solid rgba(16,185,129,0.35);
      }
      .kpi.kpi-c{
        background: linear-gradient(135deg, rgba(245,158,11,0.20), rgba(251,191,36,0.10));
        border:1px solid rgba(245,158,11,0.35);
      }

      .kpi-label{color:#e5e7eb; font-weight:800; font-size:14px; letter-spacing:0.01em;}
      .kpi-value{color:#ffffff; font-weight:900; font-size:44px; margin-top:6px; line-height:1;}
      .kpi-sub{color:#cbd5e1; font-size:12px; margin-top:6px;}

      .panel{
        background:#ffffff; border:1px solid #e5e7eb; border-radius:18px;
        padding:14px 16px; box-shadow:0 10px 25px rgba(0,0,0,0.05);
      }
      hr {border:none; border-top:1px solid #e5e7eb; margin: 16px 0;}
      .caption{color:#6b7280; font-size:12px;}
      .btnrow{display:flex; gap:10px; align-items:center; margin: 6px 0 10px 0;}
      .mapwrap{border-radius:18px; overflow:hidden;}

      /* ✅ FORZAR VISIBILIDAD DE PESTAÑAS (tabs) */
      div[data-baseweb="tab-list"]{
        display:flex !important;
        gap: 10px !important;
        padding: 10px 6px !important;
        border-bottom: 1px solid rgba(229,231,235,0.18) !important;
        overflow-x: auto !important;
      }
      button[data-baseweb="tab"]{
        color: #e5e7eb !important;
        background: rgba(255,255,255,0.06) !important;
        border-radius: 12px !important;
        padding: 10px 14px !important;
        border: 1px solid rgba(255,255,255,0.10) !important;
        font-weight: 800 !important;
        white-space: nowrap !important;
      }
      button[data-baseweb="tab"][aria-selected="true"]{
        background: rgba(99,102,241,0.20) !important;
        border: 1px solid rgba(99,102,241,0.35) !important;
      }
    </style>
    """,
    unsafe_allow_html=True
)

# =========================
# SESSION STATE
# =========================
if "map_fullscreen_tab1" not in st.session_state:
    st.session_state["map_fullscreen_tab1"] = False
if "map_fullscreen_tab2" not in st.session_state:
    st.session_state["map_fullscreen_tab2"] = False
if "map_fullscreen_tab3" not in st.session_state:
    st.session_state["map_fullscreen_tab3"] = False
if "map_fullscreen_tab4" not in st.session_state:
    st.session_state["map_fullscreen_tab4"] = False

# =========================
# COLORES POR PROVINCIA
# =========================
PROV_COLORS = {
    "San Jose":    {"stroke": "#6d28d9", "fill": "#8b5cf6"},  # morado
    "Alajuela":    {"stroke": "#0f766e", "fill": "#14b8a6"},  # teal
    "Cartago":     {"stroke": "#b45309", "fill": "#f59e0b"},  # naranja
    "Heredia":     {"stroke": "#1d4ed8", "fill": "#60a5fa"},  # azul
    "Guanacaste":  {"stroke": "#15803d", "fill": "#4ade80"},  # verde
    "Puntarenas":  {"stroke": "#b91c1c", "fill": "#f87171"},  # rojo
    "Limon":       {"stroke": "#a16207", "fill": "#facc15"},  # amarillo
}

# Orden fijo solicitado para tablas
PROV_ORDER = ["San Jose", "Alajuela", "Cartago", "Heredia", "Guanacaste", "Puntarenas", "Limon"]

def sort_by_provincia(df: pd.DataFrame) -> pd.DataFrame:
    if "provincia" not in df.columns:
        return df
    out = df.copy()
    out["__prov_ord__"] = out["provincia"].apply(lambda x: PROV_ORDER.index(x) if x in PROV_ORDER else 999)
    sort_cols = ["__prov_ord__", "provincia"]
    if "canton" in out.columns:
        sort_cols.append("canton")
    out = out.sort_values(sort_cols).drop(columns=["__prov_ord__"])
    return out

# =========================
# NORMALIZACIÓN DE TEXTO
# =========================
def clean_txt(x: str) -> str:
    if pd.isna(x):
        return ""
    x = str(x).strip()
    x = re.sub(r"\s+", " ", x)
    return x

def normalize_estructura(name: str) -> str:
    name = clean_txt(name)
    if not name:
        return ""
    low = name.lower()
    if low.startswith("diablo"):
        return "Diablo - Alejandro Arias Monge"
    return name

def title_case_headers(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [c[:1].upper() + c[1:] if isinstance(c, str) and c else c for c in df.columns]
    return df

# =========================
# COORDENADAS (centroides cantonales aprox)
# =========================
CANTON_COORDS = {
    # San José
    "San Jose": (9.9326, -84.0790),
    "Escazu": (9.9189, -84.1386),
    "Desamparados": (9.8985, -84.0664),
    "Puriscal": (9.8490, -84.3160),
    "Tarrazu": (9.6580, -84.0130),
    "Aserri": (9.8585, -84.0928),
    "Mora": (9.9095, -84.3405),
    "Goicoechea": (9.9527, -84.0595),
    "Santa Ana": (9.9300, -84.1820),
    "Alajuelita": (9.9060, -84.1050),
    "Vasquez de Coronado": (9.9760, -84.0100),
    "Acosta": (9.8020, -84.3400),
    "Tibas": (9.9560, -84.0800),
    "Moravia": (9.9610, -84.0470),
    "Montes de Oca": (9.9400, -84.0400),
    "Turrubares": (9.9070, -84.4800),
    "Dota": (9.6480, -83.9890),
    "Curridabat": (9.9130, -84.0320),
    "Perez Zeledon": (9.3720, -83.7030),
    "Leon Cortes": (9.6770, -84.0470),

    # Alajuela
    "Alajuela": (10.0163, -84.2147),
    "San Ramon": (10.0880, -84.4700),
    "Grecia": (10.0730, -84.3140),
    "San Mateo": (9.9500, -84.5240),
    "Atenas": (9.9800, -84.3820),
    "Naranjo": (10.0970, -84.3780),
    "Palmares": (10.0600, -84.4330),
    "Poas": (10.1000, -84.2330),
    "Orotina": (9.9100, -84.5220),
    "San Carlos": (10.3300, -84.4300),
    "Zarcero": (10.1860, -84.3920),
    "Sarchi": (10.0930, -84.3440),
    "Upala": (10.9000, -85.0200),
    "Los Chiles": (11.0300, -84.7200),
    "Guatuso": (10.7100, -84.9300),
    "Rio Cuarto": (10.3350, -84.2000),

    # Cartago
    "Cartago": (9.8644, -83.9194),
    "Paraiso": (9.8380, -83.8660),
    "La Union": (9.9190, -83.9890),
    "Jimenez": (9.9790, -83.7100),
    "Turrialba": (9.9040, -83.6840),
    "Alvarado": (9.9800, -83.8000),
    "Oreamuno": (9.8900, -83.8600),
    "El Guarco": (9.8000, -83.9100),

    # Heredia
    "Heredia": (9.9980, -84.1160),
    "Barba": (10.0200, -84.1200),
    "Santo Domingo": (9.9800, -84.0900),
    "Santa Barbara": (10.0400, -84.1500),
    "San Rafael": (10.0500, -84.1100),
    "San Isidro": (10.0600, -84.0900),
    "Belen": (9.9800, -84.1900),
    "Flores": (10.0000, -84.1600),
    "San Pablo": (10.0000, -84.1000),
    "Sarapiqui": (10.4600, -84.0300),

    # Guanacaste
    "Liberia": (10.6340, -85.4370),
    "Nicoya": (10.1490, -85.4520),
    "Santa Cruz": (10.2600, -85.5860),
    "Bagaces": (10.5300, -85.2500),
    "Carrillo": (10.4300, -85.5500),
    "Cañas": (10.4300, -85.1000),
    "Abangares": (10.2200, -84.9000),
    "Tilaran": (10.4700, -84.9700),
    "Nandayure": (9.9200, -85.2900),
    "La Cruz": (11.0700, -85.6300),
    "Hojancha": (10.0600, -85.4200),

    # Puntarenas
    "Puntarenas": (9.9800, -84.8300),
    "Esparza": (9.9900, -84.6600),
    "Buenos Aires": (9.1700, -83.3300),
    "Montes de Oro": (10.1000, -84.7300),
    "Osa": (8.9500, -83.5300),
    "Aguirre": (9.4300, -84.1600),
    "Golfito": (8.6500, -83.1500),
    "Coto Brus": (8.9500, -82.9500),
    "Parrita": (9.5200, -84.3200),
    "Corredores": (8.5800, -82.9500),
    "Garabito": (9.6200, -84.6300),
    "Monteverde": (10.3000, -84.8200),
    "Puerto Jimenez": (8.5300, -83.3000),
    "Quepos": (9.4320, -84.1620),

    # Limón
    "Limon": (9.9900, -83.0300),
    "Pococi": (10.6200, -83.7400),
    "Siquirres": (10.1000, -83.5100),
    "Talamanca": (9.6200, -82.8500),
    "Matina": (10.0800, -83.3000),
    "Guacimo": (10.2100, -83.6800),
}

# =========================
# CANTON -> PROVINCIA (para datasets sin provincia)
# =========================
CANTON_TO_PROV = {
    # San Jose
    "San Jose": "San Jose",
    "Acosta": "San Jose",
    "Aserri": "San Jose",
    "Alajuelita": "San Jose",
    "Desamparados": "San Jose",
    "Dota": "San Jose",
    "Curridabat": "San Jose",
    "Goicoechea": "San Jose",
    "Montes de Oca": "San Jose",
    "Mora": "San Jose",
    "Puriscal": "San Jose",
    "Santa Ana": "San Jose",
    "Tarraszu": "San Jose",
    "Tarrazu": "San Jose",
    "Tibás": "San Jose",
    "Tibas": "San Jose",
    "Turrubares": "San Jose",
    "Vasquez de Coronado": "San Jose",
    "Perez Zeledon": "San Jose",
    "Leon Cortes": "San Jose",

    # Alajuela
    "Alajuela": "Alajuela",
    "Atenas": "Alajuela",
    "Grecia": "Alajuela",
    "Guatuso": "Alajuela",
    "Los Chiles": "Alajuela",
    "Naranjo": "Alajuela",
    "Orotina": "Alajuela",
    "Palmares": "Alajuela",
    "Peñas Blancas": "Alajuela",
    "Rio Cuarto": "Alajuela",
    "San Carlos": "Alajuela",
    "San Mateo": "Alajuela",
    "San Ramon": "Alajuela",
    "Upala": "Alajuela",
    "Zarcero": "Alajuela",
    "Sarchi": "Alajuela",
    "Poas": "Alajuela",

    # Cartago
    "Cartago": "Cartago",
    "Cervantes": "Cartago",
    "Jimenez": "Cartago",
    "La Union": "Cartago",
    "Tucurrique": "Cartago",
    "Turrialba": "Cartago",
    "Paraiso": "Cartago",
    "Oreamuno": "Cartago",
    "Alvarado": "Cartago",
    "El Guarco": "Cartago",

    # Heredia
    "Heredia": "Heredia",
    "Barva": "Heredia",
    "Flores": "Heredia",
    "San Pablo": "Heredia",
    "Santa Barbara": "Heredia",
    "Sarapiqui": "Heredia",

    # Guanacaste
    "Abangares": "Guanacaste",
    "Bagaces": "Guanacaste",
    "Colorado": "Guanacaste",
    "La Cruz": "Guanacaste",
    "Nicoya": "Guanacaste",
    "Santa Cruz": "Guanacaste",
    "Tilaran": "Guanacaste",
    "Liberia": "Guanacaste",

    # Puntarenas
    "Cóbano": "Puntarenas",
    "Cobano": "Puntarenas",
    "Coto Brus": "Puntarenas",
    "Puntarenas": "Puntarenas",
    "Quepos": "Puntarenas",
    "Osa": "Puntarenas",
    "Corredores": "Puntarenas",

    # Limon
    "Limon": "Limon",
    "Limón": "Limon",
    "Matina": "Limon",
    "Pococi": "Limon",
    "Pococí": "Limon",
    "Siquirres": "Limon",
}

def extract_main_canton(raw: str) -> str:
    """
    Ejemplos:
      "Alajuela / Los Cocos" -> "Alajuela"
      "Desamparados/ Los Guido" -> "Desamparados"
      "La Cruz Santa Cecilia" -> "La Cruz"
      "Limón / Limoncito" -> "Limon"
    """
    s = clean_txt(raw)
    if not s:
        return ""
    s = s.replace("Limón", "Limon")
    # Si hay "/", tomar lo de la izquierda
    if "/" in s:
        left = s.split("/")[0].strip()
        return left
    # Casos con "La Cruz Santa Cecilia"
    if s.lower().startswith("la cruz"):
        return "La Cruz"
    # Casos "San José / Pavas"
    if s.lower().startswith("san jos"):
        return "San Jose"
    # Caso "Alajuela y nombre extenso" (de tu nota) => solo Alajuela
    if s.lower().startswith("alajuela"):
        return "Alajuela"
    return s

def add_coords(df: pd.DataFrame, canton_col: str = "canton") -> pd.DataFrame:
    df = df.copy()
    df["lat"] = df[canton_col].map(lambda c: CANTON_COORDS.get(c, (None, None))[0])
    df["lon"] = df[canton_col].map(lambda c: CANTON_COORDS.get(c, (None, None))[1])
    return df

def ensure_provincia(df: pd.DataFrame, canton_col: str = "canton") -> pd.DataFrame:
    df = df.copy()
    if "provincia" not in df.columns:
        df["provincia"] = df[canton_col].map(lambda c: CANTON_TO_PROV.get(c, ""))
    return df

# =========================
# EXPORT EXCEL BONITO (sin coordenadas)
# =========================
def df_to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Datos") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    # Copia sin lat/lon
    out = df.copy()
    for col in ["lat", "lon"]:
        if col in out.columns:
            out = out.drop(columns=[col])

    # Titulos con primera letra mayuscula
    out = title_case_headers(out)

    headers = list(out.columns)
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F2937")  # gris oscuro
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="9CA3AF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    # Filas
    for _, row in out.iterrows():
        ws.append(list(row.values))

    # Estilo general
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    # Anchos automáticos
    for col_idx in range(1, ws.max_column + 1):
        max_len = 10
        col_letter = get_column_letter(col_idx)
        for r in range(1, ws.max_row + 1):
            val = ws.cell(row=r, column=col_idx).value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(60, max_len + 2)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
# =========================
# DATASET 1: ESTRUCTURAS (TU DATA ORIGINAL)
# =========================
RAW_BY_PROV = {
    "San Jose": [
        ("San Jose", [
            "Los Lara (San Sebastia)", "Los coqueros (Pavas)", "Los Moreco", "Turesky", "Pollo",
            "Indio", "Ojos Bellos", "Los Picudos (Carpio)", "Los Diablos (Pavas)", "Los Polacos (Pavas)"
        ]),
        ("Escazu", ["Los Lara", "", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "", "", ""]),
        ("Desamparados", ["Los Lara", "", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "", "", ""]),
        ("Puriscal", ["Los Lara", "", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "", "", ""]),
        ("Tarrazu", ["Los Lara", "", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "", "", ""]),
        ("Aserri", ["Los Lara", "", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "", "", ""]),
        ("Mora", ["Los Lara", "", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "", "", ""]),
        ("Goicoechea", ["Los Lara", "Mongo", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "", "", ""]),
        ("Santa Ana", ["Los Lara", "La H", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "", "", ""]),
        ("Alajuelita", ["Los Lara", "", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "", "", ""]),
        ("Vasquez de Coronado", ["", "", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "", "", ""]),
        ("Acosta", ["", "", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "", "", ""]),
        ("Tibas", ["Los Lara", "", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "", "", ""]),
        ("Moravia", ["", "", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "", "", ""]),
        ("Montes de Oca", ["Los Lara", "Cartel de Sinaloa", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "", "", ""]),
        ("Turrubares", ["", "", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "", "", ""]),
        ("Dota", ["", "", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "", "", ""]),
        ("Curridabat", ["GaryGery", "Churro y Tauro", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "", "", ""]),
        ("Perez Zeledon", ["", "", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "", "", ""]),
        ("Leon Cortes", ["", "", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "", "", ""]),
    ],
    "Alajuela": [
        ("Alajuela", ["La hyena", "Diablo - Alejandro Arias", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "Los Ungas", "Cascaritas (San Antonio y El Roble)", ""]),
        ("San Ramon", ["", "", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "", "", ""]),
        ("Grecia", ["", "", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "", "", ""]),
        ("San Mateo", ["", "", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "", "", ""]),
        ("Atenas", ["", "", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "", "", ""]),
        ("Naranjo", ["", "", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "", "", ""]),
        ("Palmares", ["", "", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "", "", ""]),
        ("Poas", ["", "", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "", "", ""]),
        ("Orotina", ["", "", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "", "", ""]),
        ("San Carlos", ["", "", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "Diablo", "", ""]),
        ("Zarcero", ["", "", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "", "", ""]),
        ("Sarchi", ["", "", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "Gordon", "", ""]),
        ("Upala", ["", "", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "", "", ""]),
        ("Los Chiles", ["", "", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "", "", ""]),
        ("Guatuso", ["", "", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "", "", ""]),
        ("Rio Cuarto", ["", "", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "", "", ""]),
    ],
    "Cartago": [
        ("Cartago", ["Los Maruja", "", "", "", "", "", "", "Chacales", "Pollo", "Turco"]),
        ("Paraiso", ["Los Maruja", "", "", "", "", "", "", "", "", ""]),
        ("La Union", ["Los Maruja", "", "", "", "", "", "", "Hermanos Gary Gery", "", ""]),
        ("Jimenez", ["Los Maruja", "", "", "", "", "", "", "", "", ""]),
        ("Turrialba", ["Los Maruja", "", "", "", "", "", "", "Diablo", "", ""]),
        ("Alvarado", ["Los Maruja", "", "", "", "", "", "", "", "", ""]),
        ("Oreamuno", ["Los Maruja", "", "", "", "", "", "", "Los Elizondo", "", ""]),
        ("El Guarco", ["Los Maruja", "", "", "", "", "", "", "Hermanos Gary Gery", "Palomo", "Los Maruja"]),
    ],
    "Heredia": [
        ("Heredia", ["Lara", "Myrie", "Polacos", "Hermanos Ga", "Shaggy", "", "", "Pipis (Guararri y Los ...)", "Zepol", ""]),
        ("Barba", ["", "", "", "", "", "", "", "", "", ""]),
        ("Santo Domingo", ["", "", "", "", "", "", "", "", "", ""]),
        ("Santa Barbara", ["", "", "", "", "", "", "", "", "", ""]),
        ("San Rafael", ["", "", "", "", "", "", "", "", "", ""]),
        ("San Isidro", ["", "", "", "", "", "", "", "", "", ""]),
        ("Belen", ["", "", "", "", "", "", "", "", "", ""]),
        ("Flores", ["", "", "", "", "", "", "", "", "", ""]),
        ("San Pablo", ["Diablo - Alejandro Arias Monge", "", "", "", "", "", "", "", "", ""]),
        ("Sarapiqui", ["Diablo - Alejandro Arias Monge", "", "", "", "", "", "", "Diablo", "", ""]),
    ],
    "Guanacaste": [
        ("Liberia",   ["Diablo - Alejandro Arias Monge", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "Cartel de Juárez", "", "", ""]),
        ("Nicoya",    ["Diablo - Alejandro Arias Monge", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "Cartel de Juárez", "", "", ""]),
        ("Santa Cruz",["Diablo - Alejandro Arias Monge", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "Cartel de Juárez", "Diablo", "Ticachu (Tamarindo)", "Los Porteños"]),
        ("Bagaces",   ["Diablo - Alejandro Arias Monge", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "Cartel de Juárez", "", "", ""]),
        ("Carrillo",  ["Diablo - Alejandro Arias Monge", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "Cartel de Juárez", "", "", ""]),
        ("Cañas",     ["Diablo - Alejandro Arias Monge", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "Cartel de Juárez", "Rata", "", ""]),
        ("Abangares", ["Diablo - Alejandro Arias Monge", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "Cartel de Juárez", "", "", ""]),
        ("Tilaran",   ["Diablo - Alejandro Arias Monge", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "Cartel de Juárez", "", "", ""]),
        ("Nandayure", ["Diablo - Alejandro Arias Monge", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "Cartel de Juárez", "", "", ""]),
        ("La Cruz",   ["Diablo - Alejandro Arias Monge", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "Cartel de Juárez", "", "", ""]),
        ("Hojancha",  ["Diablo - Alejandro Arias Monge", "Los Moreco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "Cartel de Juárez", "", "", ""]),
    ],
    "Puntarenas": [
        ("Puntarenas", ["Diablo - Alejandro Arias Monge", "Los Picachu-20 de No", "Guayacanes - El Gordo Dan", "Los Buhos -", "El Gordo Ram", "Los Leiner - Barranca", "Los Unga - Barranca", "", "", ""]),
        ("Esparza",      ["Diablo - Alejandro Arias Monge", "Cartel de Juárez", "Los Unga - B", "El Gordo Ramos - Barranca", "", "", "", "", "", ""]),
        ("Buenos Aires", ["Diablo - Alejandro Arias Monge", "Cartel de Juárez", "", "", "", "", "", "", "", ""]),
        ("Montes de Oro",["Diablo - Alejandro Arias Monge", "Cartel de Juárez", "", "", "", "", "", "", "", ""]),
        ("Osa",          ["Diablo - Alejandro Arias Monge", "Cartel de Juárez", "", "", "", "", "", "Diablo", "", ""]),
        ("Aguirre",      ["Diablo - Alejandro Arias Monge", "Cartel de Juárez", "", "", "", "", "", "", "", ""]),
        ("Golfito",      ["Diablo - Alejandro Arias Monge", "Cartel de Juárez", "", "", "", "", "", "", "", ""]),
        ("Coto Brus",    ["Diablo - Alejandro Arias Monge", "Cartel de Juárez", "", "", "", "", "", "", "", ""]),
        ("Parrita",      ["Diablo - Alejandro Arias Monge", "Cartel de Juárez", "", "", "", "", "", "", "", ""]),
        ("Corredores",   ["Diablo - Alejandro Arias Monge", "Cartel de Juárez", "", "", "", "", "", "", "", ""]),
        ("Garabito",     ["Diablo - Alejandro Arias Monge", "Cartel de Juárez", "", "", "", "", "", "", "", ""]),
        ("Monteverde",   ["Diablo - Alejandro Arias Monge", "Cartel de Juárez", "", "", "", "", "", "", "", ""]),
        ("Puerto Jimenez",["Diablo - Alejandro Arias Monge", "Cartel de Juárez", "", "", "", "", "", "", "", ""]),
    ],
    "Limon": [
        ("Limon",     ["La H", "Diablo - Alejandro Arias Monge", "Los Morenco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "Tony Peña Russel", "La H", ""]),
        ("Pococi",    ["La H", "Diablo - Alejandro Arias Monge", "Los Morenco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "", "", ""]),
        ("Siquirres", ["La H", "Diablo - Alejandro Arias Monge", "Los Morenco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "", "", ""]),
        ("Talamanca", ["La H", "Diablo - Alejandro Arias Monge", "Los Morenco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "", "", ""]),
        ("Matina",    ["La H", "Diablo - Alejandro Arias Monge", "Los Morenco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "Tony Peña Russel", "La H", ""]),
        ("Guacimo",   ["La H", "Diablo - Alejandro Arias Monge", "Los Morenco", "Turesky", "Pollo", "Indio", "Ojos Bellos", "Pechuga", "", ""]),
    ],
}

def build_wide_df() -> pd.DataFrame:
    rows = []
    for prov, items in RAW_BY_PROV.items():
        for canton, structs in items:
            r = {"provincia": prov, "canton": canton}
            for i in range(10):
                r[f"e{i+1}"] = structs[i] if i < len(structs) else ""
            rows.append(r)
    return pd.DataFrame(rows)

def normalize_long(df_wide: pd.DataFrame) -> pd.DataFrame:
    struct_cols = [c for c in df_wide.columns if c.startswith("e")]
    long = df_wide.melt(
        id_vars=["provincia", "canton"],
        value_vars=struct_cols,
        var_name="col",
        value_name="estructura"
    )
    long["estructura"] = long["estructura"].apply(clean_txt)
    long["estructura"] = long["estructura"].apply(normalize_estructura)
    long = long[long["estructura"].str.len() > 0].copy()
    long.drop(columns=["col"], inplace=True)
    return long.reset_index(drop=True)

wide_estruct = build_wide_df()
df_estruct = add_coords(normalize_long(wide_estruct))

# =========================
# DATASET 2: BANDAS MUSICALES (provincia, canton, nombre, beneficiarios)
# =========================
BANDAS_ROWS = [
    # Alajuela
    ("Alajuela", "Alajuela", "Banda Municipal de Alajuela", 82),
    ("Alajuela", "Atenas", "Banda Escuela Municipal de Atenas", 50),
    ("Alajuela", "Grecia", "Centro de la Cultura (Academia Edol)", 10),
    ("Alajuela", "Guatuso", "Centro preventivo", 25),
    ("Alajuela", "Guatuso", "Banda Escuela Municipal de Guatuso", 25),
    ("Alajuela", "Los Chiles", "Banda Municipal de Los Chiles", 40),
    ("Alajuela", "Naranjo", "Banda Escuela de Concepción, Colegio La Candelaria, Banda Municipal", 155),
    ("Alajuela", "Orotina", "Banda Comunal de Orotina", 250),
    ("Alajuela", "Palmares", "Banda Escuela Municipal de Palmares", 178),
    ("Alajuela", "Peñas Blancas", "Banda Municipal de Peñas Blancas", 40),
    ("Alajuela", "Rio Cuarto", "Banda Comunal de Rio Cuarto", 50),
    ("Alajuela", "San Carlos", "Banda Municipal de San Carlos", 150),
    ("Alajuela", "San Mateo", "Banda Municipal de San Mateo", 100),
    ("Alajuela", "Upala", "Escuela Música Municipal de Upala", 100),
    ("Alajuela", "Zarcero", "Banda de marcha de Zarcero", 300),

    # Cartago
    ("Cartago", "Cartago", "Banda Escuela de Cartago", None),
    ("Cartago", "Cervantes", "Banda Escuela for Cerva de Cervantes", None),
    ("Cartago", "Jimenez", "Escuela Municipal de Música de Jiménez", 79),
    ("Cartago", "La Union", "Escuela de la Música de La Unión", 20),
    ("Cartago", "La Union", "Banda Municipal de Marca de la Unión", 30),
    ("Cartago", "Tucurrique", "Banda Escuela de Tucurrique", None),
    ("Cartago", "Turrialba", "Banda Municipal de Turrialba", 42),

    # Guanacaste
    ("Guanacaste", "Abangares", "Banda Melorítmica Lourdes y Pozo Azul (LP)", 32),
    ("Guanacaste", "Bagaces", "Banda Comunitaria", 27),
    ("Guanacaste", "Colorado", "Banda Municipal de Marcha de Colorado", 149),
    ("Guanacaste", "La Cruz", "Escuela de Musica La Cruz", 350),
    ("Guanacaste", "Nicoya", "Banda de marcha municipal", 180),
    ("Guanacaste", "Santa Cruz", "Centro Civico por la Paz Santa Cruz/Proy Policiamiento Comunitario", None),
    ("Guanacaste", "Tilaran", "Banda Escuela Municipal de Tilaran", 150),

    # Heredia
    ("Heredia", "Barva", "Casa de la música de Barva", 50),
    ("Heredia", "Flores", "Banda Escuela Municipal de Flores", 50),
    ("Heredia", "Heredia", "Banda Escuela for Heredia de Heredia", 100),
    ("Heredia", "San Pablo", "Escuela de Musica", 100),
    ("Heredia", "San Pablo", "Banda estudiantil MAVISA", 150),
    ("Heredia", "Santa Barbara", "Banda Municipal de Santa Bárbara", 50),
    ("Heredia", "Sarapiqui", "Banda Escuela Piano de Sarapiqui", 25),
    ("Heredia", "Sarapiqui", "Banda Escuela de Sarapiqui", 100),

    # (Institucion)
    ("San Jose", "Vice Paz", "Centro Civico por la Paz", 25),

    # Limón
    ("Limon", "Limon", "Kawe Calipso Youth", 30),
    ("Limon", "Matina", "Banda Municipal d Matina", 45),
    ("Limon", "Pococi", "Banda Escuela Municipal de Pococi", 100),
    ("Limon", "Pococi", "Banda Cospnli de for NPA Pococi", 100),
    ("Limon", "Siquirres", "Banda Escuela for Siqui de Siquirres", 50),

    # Puntarenas
    ("Puntarenas", "Cóbano", "Banda Escuela for CMD C de Cóbano", 75),
    ("Puntarenas", "Coto Brus", "Banda Municipal de Coto Brus", 75),
    ("Puntarenas", "Coto Brus", "Colideportivo de Brus", 50),
    ("Puntarenas", "Puntarenas", "Banda de Puntarenas Barranca", 150),
    ("Puntarenas", "Quepos", "Banda Escuela for Munic de Quepos", 120),

    # San José
    ("San Jose", "Acosta", "Banda Escuela Instrumen de Acosta", 50),
    ("San Jose", "Aserri", "Banda Escuela for Aserr de Aserri", 50),
    ("San Jose", "Desamparados", "Banda Municipal de Desamparados", 195),
    ("San Jose", "Desamparados", "Banda Municipal de Desamparados", 0),
    ("San Jose", "Desamparados", "Banda Escuela de Desamparados", 75),
    ("San Jose", "Dota", "Banda Escuela for Dota", 100),
    ("San Jose", "Montes de Oca", "Banda Estudiantil Liceo Anastasio Alfaro", 100),
    ("San Jose", "Montes de Oca", "Banda Estudiantil Liceo Anastasio Alfaro", 0),
    ("San Jose", "Puriscal", "Banda Municipal de Puriscal", 130),
    ("San Jose", "San Jose", "Banda Municipal de San José", 200),
    ("San Jose", "Goicoechea", "Banda de Marcha de Heredia", 50),
    ("San Jose", "Santa Ana", "Banda Cantonal Municipal de Santa Ana", 115),
    ("San Jose", "Santa Ana", "EMAI", 100),
    ("San Jose", "Tarrazu", "Banda Escuela Municipal de Tarrazu", 30),
    ("San Jose", "Turrubares", "Banda Escuela Municipal de Turrubares", 25),
]

df_bandas = pd.DataFrame(BANDAS_ROWS, columns=["provincia", "canton", "nombre", "beneficiarios"])
df_bandas["nombre"] = df_bandas["nombre"].apply(clean_txt)
df_bandas["canton"] = df_bandas["canton"].apply(lambda x: clean_txt(x).replace("Limón", "Limon"))
df_bandas = add_coords(df_bandas)

# Beneficiarios a numérico (None -> 0 para sumar)
df_bandas["beneficiarios_num"] = pd.to_numeric(df_bandas["beneficiarios"], errors="coerce").fillna(0).astype(int)

# =========================
# GENERIC MAP RENDERER
# =========================
def render_map_generic(df_points: pd.DataFrame, height_px: int, popup_mode: str):
    fm = df_points.dropna(subset=["lat", "lon"]).copy()
    if fm.empty:
        st.warning("No hay puntos con coordenadas para mostrar.")
        return

    center_lat = float(fm["lat"].mean())
    center_lon = float(fm["lon"].mean())

    m = folium.Map(location=[center_lat, center_lon], zoom_start=8, control_scale=True, tiles=None)

    folium.TileLayer(
        tiles="https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}",
        attr="Esri, Maxar, Earthstar Geographics, CNES/Airbus DS, USDA, USGS",
        name="ESRI Satélite",
        overlay=False,
        control=True
    ).add_to(m)

    folium.TileLayer(
        tiles="https://services.arcgisonline.com/ArcGIS/rest/services/Reference/World_Boundaries_and_Places/MapServer/tile/{z}/{y}/{x}",
        attr="Esri",
        name="Límites y lugares",
        overlay=True,
        control=True,
        opacity=0.95
    ).add_to(m)

    folium.LayerControl(collapsed=True).add_to(m)

    for _, r in fm.iterrows():
        provincia = r.get("provincia", "")
        canton = r.get("canton", "")
        colors = PROV_COLORS.get(provincia, {"stroke": "#111827", "fill": "#9ca3af"})

        if popup_mode == "estructuras":
            estructuras_list = r.get("estructuras_list", [])
            tooltip = f"{canton} ({provincia}) | {len(estructuras_list)} estructuras"
            html = f"""
            <div style="font-family: Arial; font-size: 13px; line-height: 1.25;">
              <div style="font-size: 14px;"><b>{canton}</b></div>
              <div><b>Provincia:</b> {provincia}</div>
              <div><b>Estructuras:</b> {len(estructuras_list)}</div>
              <div style="margin-top:6px;"><b>Listado:</b></div>
              <ul style="margin: 6px 0 0 18px; padding: 0;">
                {''.join([f'<li>{e}</li>' for e in estructuras_list])}
              </ul>
            </div>
            """
            radius = 6 + min(16, len(estructuras_list) * 1.6)

        elif popup_mode == "bandas":
            nombre = r.get("nombre", "")
            benef = int(r.get("beneficiarios_num", 0))
            tooltip = f"{canton} ({provincia}) | {benef} beneficiarios"
            html = f"""
            <div style="font-family: Arial; font-size: 13px; line-height: 1.25;">
              <div style="font-size: 14px;"><b>{canton}</b></div>
              <div><b>Provincia:</b> {provincia}</div>
              <div style="margin-top:6px;"><b>Banda/Club:</b> {nombre}</div>
              <div><b>Beneficiarios:</b> {benef}</div>
            </div>
            """
            radius = 6 + min(18, max(1, benef) * 0.05)

        elif popup_mode == "centros":
            centro = r.get("centro", "")
            benef = int(r.get("beneficiarios", 0))
            tooltip = f"{canton} ({provincia}) | {benef} beneficiarios"
            html = f"""
            <div style="font-family: Arial; font-size: 13px; line-height: 1.25;">
              <div style="font-size: 14px;"><b>{canton}</b></div>
              <div><b>Provincia:</b> {provincia}</div>
              <div style="margin-top:6px;"><b>Centro:</b> {centro}</div>
              <div><b>Beneficiarios:</b> {benef}</div>
            </div>
            """
            radius = 6 + min(18, max(1, benef) * 0.06)

        else:  # empleabilidad
            cursos = r.get("cursos", "")
            matric = int(r.get("matriculadas", 0))
            egres = r.get("egresadas_num", r.get("egresadas", 0))
            h = int(r.get("hombres", 0))
            mcount = int(r.get("mujeres", 0))
            tooltip = f"{canton} ({provincia}) | Mat: {matric} | Egr: {egres}"
            html = f"""
            <div style="font-family: Arial; font-size: 13px; line-height: 1.25;">
              <div style="font-size: 14px;"><b>{canton}</b></div>
              <div><b>Provincia:</b> {provincia}</div>
              <div style="margin-top:6px;"><b>Matriculadas:</b> {matric}</div>
              <div><b>Egresadas:</b> {egres}</div>
              <div><b>Hombres:</b> {h} &nbsp;&nbsp; <b>Mujeres:</b> {mcount}</div>
              <div style="margin-top:6px;"><b>Cursos:</b></div>
              <div style="white-space: pre-wrap;">{cursos}</div>
            </div>
            """
            radius = 6 + min(18, max(1, matric) * 0.05)

        popup = folium.Popup(html, max_width=460)

        folium.CircleMarker(
            location=[float(r["lat"]), float(r["lon"])],
            radius=radius,
            weight=2,
            color=colors["stroke"],
            fill=True,
            fill_color=colors["fill"],
            fill_opacity=0.60,
            tooltip=tooltip,
            popup=popup
        ).add_to(m)

    st.markdown("<div class='mapwrap'>", unsafe_allow_html=True)
    st_folium(m, use_container_width=True, height=height_px)
    st.markdown("</div>", unsafe_allow_html=True)
# =========================
# DATASET 3: CENTROS PREVENTIVOS COMUNITARIOS (sin provincia -> se asigna)
# =========================
CENTROS_ROWS = [
    (40,  "Alajuela / Vistas de Santamaria", "Centro Preventivo Comunitario"),
    (25,  "Alajuela / Los Cocos", "Community Prevention Center"),
    (60,  "Alajuelita / San Felipe", "Safe Space"),
    (40,  "Alajuelita / Centro", "Casa de Creación Juvenil"),
    (40,  "Corredores/ Sabalito", "Safe Space"),
    (40,  "Coto Brus", "Safe Space"),
    (100, "Curridabat", "Human Center of Development La Cometa"),
    (50,  "Desamparados", "Civic Center of Peace & Desamparados Municipality"),
    (350, "Desamparados/ Los Guido", "Safe Space"),
    (60,  "Guatuso", "Safe Space"),
    (100, "Heredia", "Community Prevention Center"),
    (100, "La Cruz Santa Cecilia", "Safe Space"),
    (80,  "La Cruz/ Centro", "Safe Space"),
    (30,  "Limon/ Pueblo Nuevo", "Community Prevention Center: Youth Center"),
    (150, "Limon/ Cieneguita", "Community Prevention Center/ Safe Space/ Surf Boxing"),
    (15,  "Limón/ Limoncito", "Community Prevention Center/ Safe Space"),
    (30,  "Limón / Valle de la Estrella", "Colectivo deportivo Valle de la Estrella"),
    (100, "Limon/ Cocos", "Asoc de Futbol Club Atlético Limonense"),
    (150, "Los Chiles / Muelle", "Centro Preventivo Comunitario"),
    (100, "Los Chiles / La Virgen", "Safe Space"),
    (50,  "Matina/ Estrada", "Safe Space"),
    (55,  "Matina / Luzon", "Safe Space"),
    (75,  "Montes de Oca", "Community Prevention Center /Circo Social de Sinaí"),
    (60,  "Mora", "Casa de la Juventud/Club House"),
    (45,  "Osa/ Bahia Ballena", "Centro Preventivo Comunitario"),
    (20,  "Pococi/ la Sole", "ADI Sole, DINADECO, UNICEF, PANI, INL"),
    (100, "Puntarenas/ Barranca", "Centro Preventivo Comunitario  ONG Barranca Sport Club"),
    (15,  "Puntarenas/ Chacarita", "Safe Space"),
    (50,  "Puntarenas/ Fray Casiano", "Safe Space"),
    (20,  "Quepos /Pies Mojados", "Safe Space"),
    (70,  "San Carlos", "Civic Center of Peace"),
    (40,  "San José/ Pavas", "Safe Space"),
    (30,  "San José / Hatillo", "Safe Space"),
    (150, "San José / Carpio", "Safe Space"),
    (25,  "San Ramón", "Safe Space"),
    (100, "Santa Ana", "Casita de Escucha Corazón de Jesús"),
    (100, "Santa Ana", "Casita de Escucha El Triunfo"),
    (40,  "Sarapiquí / llanuras de Gaspar", "Safe Space"),
    (40,  "Sarapiquí / puerto Viejo", "Safe Space"),
    (15,  "Siquirres/ 3 Cercas", "Safe Space"),
    (25,  "Turrialba", "Community Prevention Center"),
    (52,  "Turrubares", "Safe Space"),
    (20,  "Upala/Mexico", "Safe Space"),
    (30,  "Upala/ La Real", "Safe Space"),
]

df_centros = pd.DataFrame(CENTROS_ROWS, columns=["beneficiarios", "canton_raw", "centro"])
df_centros["canton"] = df_centros["canton_raw"].apply(extract_main_canton)
df_centros = ensure_provincia(df_centros, canton_col="canton")
df_centros = add_coords(df_centros, canton_col="canton")

df_centros["beneficiarios"] = pd.to_numeric(df_centros["beneficiarios"], errors="coerce").fillna(0).astype(int)

# =========================
# DATASET 4: PROGRAMAS DE EMPLEABILIDAD (sin provincia -> se asigna)
# =========================
EMPLEA_ROWS = [
    ("Cartago", "Aseo y Limpieza de Espacios Comerciales\nServicio al Cliente\nManejo de Vehículos Pesados", 100, 95, 36, 59),
    ("Cartago", "Aseo y Limpieza de Espacios Comerciales\nServicio al Cliente\nBuenas Prácticas de Manufactura BPM\nOperario de Construcción", 100, 113, 26, 87),
    ("Turrialba", "Servicio al Cliente y Ventas\nAseo y Limpieza de Espacios Comerciales\nPistero de Gasolinera", 100, 88, 61, 27),
    ("Turrialba", "Servicio al Cliente y Ventas\nAseo y Limpieza de Espacios Comerciales\nPistero de Gasolinera\nBuenas Prácticas de Manufactura BPM", 100, 71, 48, 23),
    ("Puntarenas", "Buenas Prácticas de Manufactura BPM\nAuxiliar de Bodega\nAtención y Servicio al Cliente para Salonero\nPistero de Gasolinera\nAseo y Limpieza en Hotelería", 100, 88, 27, 61),
    ("Limon", "Aseo y Limpieza Cabinas y Hoteles\nServicio al Cliente para Comercio\nAuxiliar de Bodega\nPistero de Gasolinera", 100, 89, 20, 69),
    ("Abangares", "Auxiliar de cocina\nServicio al Cliente y ventas", 30, 31, 6, 25),
    ("Pococi", "Buenas Prácticas de Manufactura con énfasis en industria alimentaria\nAuxiliar de Bodega", 100, 97, 52, 45),
    ("Desamparados", "Servicio al cliente con énfasis en ferretería\nBuenas Prácticas de Manufactura con énfasis en industria alimentaria\nAuxiliar de Bodega", 100, 83, 30, 53),
    ("La Union", "Buenas Prácticas de Manufactura con énfasis en industria alimentaria\nAuxiliar de Bodega\nAseo y Limpieza de locales comerciales", 100, 85, 36, 49),
    ("Curridabat", "Buenas Prácticas de Manufactura con énfasis en industria alimentaria\nAuxiliar de Bodega\nAseo y Limpieza de locales comerciales", 100, 73, 23, 50),
    ("Puntarenas", "Buenas Prácticas de Manufactura BPM\nAuxiliar de Bodega\nAtención y Servicio al Cliente para Salonero\nPistero de Gasolinera\nAseo y Limpieza en Hotelería", 100, 100, 25, 75),
    ("Limon", "Servicio al cliente con énfasis en ferretería\nSaloneros\nAuxiliar de Bodega", 100, 78, 26, 52),
    ("Alajuela", "Buenas Prácticas de Manufactura en Industria Alimentaria\nAuxiliar de Bodega", 100, 87, 45, 42),
    ("Upala", "Operario Agrícola\nServicio y Atención en Hotelería y Centros de Alojamiento", 100, 89, 31, 58),
    ("Oreamuno", "Buenas Prácticas de Manufactura con énfasis en industria alimentaria", 30, 31, 6, 25),
    ("Pococi", "Servicio al cliente con énfasis en Cajas\nAuxiliar de bodega\nAuxiliar de cocina\nAseo y Limpieza de locales comerciales", 100, 100, 19, 81),
    ("Mora", "Salonero-Bartender\nCuidadoras de niños\nAuxiliar de Bodega", 40, 40, 7, 33),
    ("Goicoechea", "Buenas prácticas de Manufactura en Industria Alimentaria\nServicio y Atención en Hotelería y Centros de Alojamiento", 75, 69, 0, 75),
    ("Goicoechea", "Buenas prácticas de Manufactura en Industria Alimentaria\nServicio y Atención en Hotelería y Centros de Alojamiento", 58, 0, 10, 48),
    ("Santa Cruz", "Servicios de Hotelería y Centros de Alojamiento", 50, 29, 0, 0),
    ("La Cruz", "Peón de finca, Atención en Hotelería y Centros de Alojamiento", 50, 0, 0, 0),
    ("Bagaces", "Servicios de Hotelería y Centros de Alojamiento", 50, 47, 0, 0),
]

df_empleo = pd.DataFrame(EMPLEA_ROWS, columns=["canton_raw", "cursos", "matriculadas", "egresadas", "hombres", "mujeres"])
df_empleo["canton"] = df_empleo["canton_raw"].apply(extract_main_canton)
df_empleo = ensure_provincia(df_empleo, canton_col="canton")
df_empleo = add_coords(df_empleo, canton_col="canton")

df_empleo["matriculadas"] = pd.to_numeric(df_empleo["matriculadas"], errors="coerce").fillna(0).astype(int)
df_empleo["egresadas_num"] = pd.to_numeric(df_empleo["egresadas"], errors="coerce").fillna(0).astype(int)
df_empleo["hombres"] = pd.to_numeric(df_empleo["hombres"], errors="coerce").fillna(0).astype(int)
df_empleo["mujeres"] = pd.to_numeric(df_empleo["mujeres"], errors="coerce").fillna(0).astype(int)

# =========================
# HEADER + TABS
# =========================
st.markdown("<div class='title'>CR | Mapas y Programas</div>", unsafe_allow_html=True)
st.markdown("<div class='subtitle'>Mapas satelitales ESRI, filtros por pestaña, tablas normalizadas y descargas ordenadas.</div>", unsafe_allow_html=True)

tab1, tab2, tab3, tab4 = st.tabs([
    "Estructuras criminales",
    "Bandas Musicales",
    "Centros Preventivos Comunitarios",
    "Programas de empleabilidad",
])

# =========================
# TAB 1: ESTRUCTURAS
# =========================
with tab1:
    st.markdown("<div class='panel'>", unsafe_allow_html=True)
    st.subheader("Filtros — Estructuras")

    c1, c2, c3 = st.columns([1, 1, 1], gap="large")
    provincias = sorted(df_estruct["provincia"].unique().tolist())
    cantones = sorted(df_estruct["canton"].unique().tolist())
    estructuras = sorted(df_estruct["estructura"].unique().tolist())

    with c1:
        prov_sel = st.multiselect("Provincia", provincias, default=[], key="t1_prov")
    with c2:
        cant_sel = st.multiselect("Cantón", cantones, default=[], key="t1_cant")
    with c3:
        estr_sel = st.multiselect("Estructura", estructuras, default=[], key="t1_estr")

    f1 = df_estruct.copy()
    if prov_sel:
        f1 = f1[f1["provincia"].isin(prov_sel)]
    if cant_sel:
        f1 = f1[f1["canton"].isin(cant_sel)]
    if estr_sel:
        f1 = f1[f1["estructura"].isin(estr_sel)]

    cantones_unicos = f1["canton"].nunique()
    estructuras_unicas = f1["estructura"].nunique()

    st.markdown(
        f"""
        <div class="kpi-grid">
          <div class="kpi kpi-a">
            <div class="kpi-label">Cantones</div>
            <div class="kpi-value">{cantones_unicos:,}</div>
            <div class="kpi-sub">Total según filtros</div>
          </div>
          <div class="kpi kpi-b">
            <div class="kpi-label">Estructuras</div>
            <div class="kpi-value">{estructuras_unicas:,}</div>
            <div class="kpi-sub">Únicas según filtros</div>
          </div>
        </div>
        """,
        unsafe_allow_html=True
    )

    st.markdown("</div>", unsafe_allow_html=True)
    st.markdown("<hr/>", unsafe_allow_html=True)

    # Agrupar para mapa (lista de estructuras por cantón)
    fm1 = f1.dropna(subset=["lat", "lon"]).copy()
    grp1 = (
        fm1.groupby(["provincia", "canton", "lat", "lon"])
        .agg(estructuras_list=("estructura", lambda s: sorted(set(s))))
        .reset_index()
    )

    # Fullscreen
    if st.session_state["map_fullscreen_tab1"]:
        st.markdown("<div class='btnrow'>", unsafe_allow_html=True)
        if st.button("⬅️ Salir de pantalla completa", key="t1_exit_full"):
            st.session_state["map_fullscreen_tab1"] = False
            st.rerun()
        st.markdown("<span class='caption'>Modo pantalla completa.</span>", unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

        render_map_generic(grp1, height_px=920, popup_mode="estructuras")
        st.stop()

    left, right = st.columns([1.05, 0.95], gap="large")

    with left:
        st.markdown("<div class='panel'>", unsafe_allow_html=True)
        st.subheader("Top 10 estructuras (conteos)")

        top_struct = (
            f1.groupby("estructura")
            .size()
            .reset_index(name="conteo")
            .sort_values("conteo", ascending=False)
            .head(10)
        )
        fig_bar = px.bar(top_struct, x="conteo", y="estructura", orientation="h")
        fig_bar.update_layout(height=430, margin=dict(l=10, r=10, t=10, b=10))
        st.plotly_chart(fig_bar, use_container_width=True)

        st.subheader("Tabla normalizada")
        tabla1 = (
            f1.groupby(["provincia", "canton"])["estructura"]
            .apply(lambda s: ", ".join(sorted(set(s))))
            .reset_index()
            .rename(columns={"estructura": "estructuras"})
        )
        tabla1 = sort_by_provincia(tabla1)

        st.dataframe(
            title_case_headers(tabla1),
            use_container_width=True,
            height=360,
            hide_index=True
        )

        st.markdown("</div>", unsafe_allow_html=True)

    with right:
        st.markdown("<div class='panel'>", unsafe_allow_html=True)
        st.subheader("Mapa satelital (ESRI) — puntos por cantón")

        st.markdown("<div class='btnrow'>", unsafe_allow_html=True)
        if st.button("⛶ Ver mapa en pantalla completa", key="t1_full"):
            st.session_state["map_fullscreen_tab1"] = True
            st.rerun()
        st.markdown("<span class='caption'>Abre el mapa ocupando casi toda la pantalla.</span>", unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

        render_map_generic(grp1, height_px=820, popup_mode="estructuras")
        st.markdown("</div>", unsafe_allow_html=True)

    st.markdown("<hr/>", unsafe_allow_html=True)

    # Descargas
    excel1 = df_to_excel_bytes(tabla1, sheet_name="Estructuras")
    st.download_button(
        "⬇️ Descargar Excel (Estructuras)",
        data=excel1,
        file_name="estructuras_normalizado.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_xlsx_t1"
    )

    st.markdown(
        f"<div class='caption'>Resumen: <b>{estructuras_unicas}</b> estructuras únicas en <b>{cantones_unicos}</b> cantones (según filtros).</div>",
        unsafe_allow_html=True
    )
# =========================
# TAB 2: BANDAS MUSICALES
# =========================
with tab2:
    st.markdown("<div class='panel'>", unsafe_allow_html=True)
    st.subheader("Filtros — Bandas Musicales")

    c1, c2, c3 = st.columns([1, 1, 1], gap="large")
    provincias2 = sorted(df_bandas["provincia"].unique().tolist())
    cantones2 = sorted(df_bandas["canton"].unique().tolist())
    nombres2 = sorted(df_bandas["nombre"].unique().tolist())

    with c1:
        prov_sel2 = st.multiselect("Provincia", provincias2, default=[], key="t2_prov")
    with c2:
        cant_sel2 = st.multiselect("Cantón", cantones2, default=[], key="t2_cant")
    with c3:
        name_sel2 = st.multiselect("Nombre", nombres2, default=[], key="t2_name")

    f2 = df_bandas.copy()
    if prov_sel2:
        f2 = f2[f2["provincia"].isin(prov_sel2)]
    if cant_sel2:
        f2 = f2[f2["canton"].isin(cant_sel2)]
    if name_sel2:
        f2 = f2[f2["nombre"].isin(name_sel2)]

    cantones_unicos2 = f2["canton"].nunique()
    total_benef2 = int(f2["beneficiarios_num"].sum())

    st.markdown(
        f"""
        <div class="kpi-grid">
          <div class="kpi kpi-a">
            <div class="kpi-label">Cantones</div>
            <div class="kpi-value">{cantones_unicos2:,}</div>
            <div class="kpi-sub">Total según filtros</div>
          </div>
          <div class="kpi kpi-c">
            <div class="kpi-label">Beneficiarios</div>
            <div class="kpi-value">{total_benef2:,}</div>
            <div class="kpi-sub">Suma según filtros</div>
          </div>
        </div>
        """,
        unsafe_allow_html=True
    )
    st.markdown("</div>", unsafe_allow_html=True)
    st.markdown("<hr/>", unsafe_allow_html=True)

    # Fullscreen
    if st.session_state["map_fullscreen_tab2"]:
        st.markdown("<div class='btnrow'>", unsafe_allow_html=True)
        if st.button("⬅️ Salir de pantalla completa", key="t2_exit_full"):
            st.session_state["map_fullscreen_tab2"] = False
            st.rerun()
        st.markdown("<span class='caption'>Modo pantalla completa.</span>", unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

        render_map_generic(f2.dropna(subset=["lat", "lon"]), height_px=920, popup_mode="bandas")
        st.stop()

    left, right = st.columns([1.05, 0.95], gap="large")

    with left:
        st.markdown("<div class='panel'>", unsafe_allow_html=True)
        st.subheader("Beneficiarios por provincia")

        prov_sum = (
            f2.groupby("provincia")["beneficiarios_num"]
            .sum()
            .reset_index(name="beneficiarios")
        )
        prov_sum["__ord__"] = prov_sum["provincia"].apply(lambda x: PROV_ORDER.index(x) if x in PROV_ORDER else 999)
        prov_sum = prov_sum.sort_values("__ord__").drop(columns=["__ord__"])

        fig = px.bar(prov_sum, x="provincia", y="beneficiarios")
        fig.update_layout(height=360, margin=dict(l=10, r=10, t=10, b=10))
        st.plotly_chart(fig, use_container_width=True)

        st.subheader("Tabla normalizada")
        tabla2 = (
            f2.groupby(["provincia", "canton"])
            .agg(
                beneficiarios=("beneficiarios_num", "sum"),
                nombres=("nombre", lambda s: ", ".join(sorted(set(s))))
            )
            .reset_index()
        )
        tabla2 = sort_by_provincia(tabla2)

        st.dataframe(
            title_case_headers(tabla2[["provincia", "canton", "beneficiarios", "nombres"]]),
            use_container_width=True,
            height=360,
            hide_index=True
        )
        st.markdown("</div>", unsafe_allow_html=True)

    with right:
        st.markdown("<div class='panel'>", unsafe_allow_html=True)
        st.subheader("Mapa satelital (ESRI) — bandas y beneficiarios")

        st.markdown("<div class='btnrow'>", unsafe_allow_html=True)
        if st.button("⛶ Ver mapa en pantalla completa", key="t2_full"):
            st.session_state["map_fullscreen_tab2"] = True
            st.rerun()
        st.markdown("<span class='caption'>Puntos con color por provincia, popup con cantón y beneficiarios.</span>", unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

        render_map_generic(f2.dropna(subset=["lat", "lon"]), height_px=820, popup_mode="bandas")
        st.markdown("</div>", unsafe_allow_html=True)

    st.markdown("<hr/>", unsafe_allow_html=True)

    excel2 = df_to_excel_bytes(tabla2[["provincia", "canton", "beneficiarios", "nombres"]], sheet_name="Bandas")
    st.download_button(
        "⬇️ Descargar Excel (Bandas Musicales)",
        data=excel2,
        file_name="bandas_musicales.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_xlsx_t2"
    )

# =========================
# TAB 3: CENTROS PREVENTIVOS COMUNITARIOS
# =========================
with tab3:
    st.markdown("<div class='panel'>", unsafe_allow_html=True)
    st.subheader("Filtros — Centros Preventivos Comunitarios")

    c1, c2, c3 = st.columns([1, 1, 1], gap="large")
    provincias3 = sorted(df_centros["provincia"].unique().tolist())
    cantones3 = sorted(df_centros["canton"].unique().tolist())
    centros3 = sorted(df_centros["centro"].unique().tolist())

    with c1:
        prov_sel3 = st.multiselect("Provincia", provincias3, default=[], key="t3_prov")
    with c2:
        cant_sel3 = st.multiselect("Cantón", cantones3, default=[], key="t3_cant")
    with c3:
        centro_sel3 = st.multiselect("Nombre del centro", centros3, default=[], key="t3_centro")

    f3 = df_centros.copy()
    if prov_sel3:
        f3 = f3[f3["provincia"].isin(prov_sel3)]
    if cant_sel3:
        f3 = f3[f3["canton"].isin(cant_sel3)]
    if centro_sel3:
        f3 = f3[f3["centro"].isin(centro_sel3)]

    cantones_unicos3 = f3["canton"].nunique()
    total_benef3 = int(f3["beneficiarios"].sum())

    st.markdown(
        f"""
        <div class="kpi-grid">
          <div class="kpi kpi-a">
            <div class="kpi-label">Cantones</div>
            <div class="kpi-value">{cantones_unicos3:,}</div>
            <div class="kpi-sub">Total según filtros</div>
          </div>
          <div class="kpi kpi-c">
            <div class="kpi-label">Beneficiarios</div>
            <div class="kpi-value">{total_benef3:,}</div>
            <div class="kpi-sub">Suma según filtros</div>
          </div>
        </div>
        """,
        unsafe_allow_html=True
    )
    st.markdown("</div>", unsafe_allow_html=True)
    st.markdown("<hr/>", unsafe_allow_html=True)

    # Fullscreen
    if st.session_state["map_fullscreen_tab3"]:
        st.markdown("<div class='btnrow'>", unsafe_allow_html=True)
        if st.button("⬅️ Salir de pantalla completa", key="t3_exit_full"):
            st.session_state["map_fullscreen_tab3"] = False
            st.rerun()
        st.markdown("<span class='caption'>Modo pantalla completa.</span>", unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

        render_map_generic(f3.dropna(subset=["lat", "lon"]), height_px=920, popup_mode="centros")
        st.stop()

    left, right = st.columns([1.05, 0.95], gap="large")

    with left:
        st.markdown("<div class='panel'>", unsafe_allow_html=True)
        st.subheader("Beneficiarios por provincia")

        prov_sum3 = (
            f3.groupby("provincia")["beneficiarios"]
            .sum()
            .reset_index(name="beneficiarios")
        )
        prov_sum3["__ord__"] = prov_sum3["provincia"].apply(lambda x: PROV_ORDER.index(x) if x in PROV_ORDER else 999)
        prov_sum3 = prov_sum3.sort_values("__ord__").drop(columns=["__ord__"])

        fig3 = px.bar(prov_sum3, x="provincia", y="beneficiarios")
        fig3.update_layout(height=360, margin=dict(l=10, r=10, t=10, b=10))
        st.plotly_chart(fig3, use_container_width=True)

        st.subheader("Tabla normalizada")
        tabla3 = (
            f3.groupby(["provincia", "canton"])
            .agg(
                beneficiarios=("beneficiarios", "sum"),
                centros=("centro", lambda s: ", ".join(sorted(set(s))))
            )
            .reset_index()
        )
        tabla3 = sort_by_provincia(tabla3)

        st.dataframe(
            title_case_headers(tabla3[["provincia", "canton", "beneficiarios", "centros"]]),
            use_container_width=True,
            height=360,
            hide_index=True
        )

        st.markdown("</div>", unsafe_allow_html=True)

    with right:
        st.markdown("<div class='panel'>", unsafe_allow_html=True)
        st.subheader("Mapa satelital (ESRI) — centros y beneficiarios")

        st.markdown("<div class='btnrow'>", unsafe_allow_html=True)
        if st.button("⛶ Ver mapa en pantalla completa", key="t3_full"):
            st.session_state["map_fullscreen_tab3"] = True
            st.rerun()
        st.markdown("<span class='caption'>Popup con cantón, centro y beneficiarios.</span>", unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

        render_map_generic(f3.dropna(subset=["lat", "lon"]), height_px=820, popup_mode="centros")
        st.markdown("</div>", unsafe_allow_html=True)

    st.markdown("<hr/>", unsafe_allow_html=True)

    excel3 = df_to_excel_bytes(tabla3[["provincia", "canton", "beneficiarios", "centros"]], sheet_name="Centros")
    st.download_button(
        "⬇️ Descargar Excel (Centros Preventivos)",
        data=excel3,
        file_name="centros_preventivos.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_xlsx_t3"
    )

# =========================
# TAB 4: PROGRAMAS DE EMPLEABILIDAD
# =========================
with tab4:
    st.markdown("<div class='panel'>", unsafe_allow_html=True)
    st.subheader("Filtros — Programas de empleabilidad")

    c1, c2, c3 = st.columns([1, 1, 1], gap="large")
    provincias4 = sorted(df_empleo["provincia"].unique().tolist())
    cantones4 = sorted(df_empleo["canton"].unique().tolist())
    cursos4 = sorted(df_empleo["cursos"].unique().tolist())

    with c1:
        prov_sel4 = st.multiselect("Provincia", provincias4, default=[], key="t4_prov")
    with c2:
        cant_sel4 = st.multiselect("Cantón", cantones4, default=[], key="t4_cant")
    with c3:
        curso_sel4 = st.multiselect("Cursos", cursos4, default=[], key="t4_curso")

    f4 = df_empleo.copy()
    if prov_sel4:
        f4 = f4[f4["provincia"].isin(prov_sel4)]
    if cant_sel4:
        f4 = f4[f4["canton"].isin(cant_sel4)]
    if curso_sel4:
        f4 = f4[f4["cursos"].isin(curso_sel4)]

    cantones_unicos4 = f4["canton"].nunique()
    total_mat = int(f4["matriculadas"].sum())
    total_egr = int(f4["egresadas_num"].sum())

    st.markdown(
        f"""
        <div class="kpi-grid">
          <div class="kpi kpi-a">
            <div class="kpi-label">Cantones</div>
            <div class="kpi-value">{cantones_unicos4:,}</div>
            <div class="kpi-sub">Total según filtros</div>
          </div>
          <div class="kpi kpi-b">
            <div class="kpi-label">Matriculadas</div>
            <div class="kpi-value">{total_mat:,}</div>
            <div class="kpi-sub">Suma según filtros</div>
          </div>
          <div class="kpi kpi-c">
            <div class="kpi-label">Egresadas</div>
            <div class="kpi-value">{total_egr:,}</div>
            <div class="kpi-sub">Suma según filtros</div>
          </div>
        </div>
        """,
        unsafe_allow_html=True
    )

    st.markdown("</div>", unsafe_allow_html=True)
    st.markdown("<hr/>", unsafe_allow_html=True)

    # Fullscreen
    if st.session_state["map_fullscreen_tab4"]:
        st.markdown("<div class='btnrow'>", unsafe_allow_html=True)
        if st.button("⬅️ Salir de pantalla completa", key="t4_exit_full"):
            st.session_state["map_fullscreen_tab4"] = False
            st.rerun()
        st.markdown("<span class='caption'>Modo pantalla completa.</span>", unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

        render_map_generic(f4.dropna(subset=["lat", "lon"]), height_px=920, popup_mode="empleabilidad")
        st.stop()

    left, right = st.columns([1.05, 0.95], gap="large")

    with left:
        st.markdown("<div class='panel'>", unsafe_allow_html=True)
        st.subheader("Gráfico 1 — Matriculadas vs Egresadas (por provincia)")

        prov_m = (
            f4.groupby("provincia")[["matriculadas", "egresadas_num"]]
            .sum()
            .reset_index()
        )
        prov_m["__ord__"] = prov_m["provincia"].apply(lambda x: PROV_ORDER.index(x) if x in PROV_ORDER else 999)
        prov_m = prov_m.sort_values("__ord__").drop(columns=["__ord__"])

        prov_m_long = prov_m.melt(id_vars=["provincia"], value_vars=["matriculadas", "egresadas_num"],
                                 var_name="tipo", value_name="cantidad")
        prov_m_long["tipo"] = prov_m_long["tipo"].replace({"egresadas_num": "egresadas"})

        fig4a = px.bar(prov_m_long, x="provincia", y="cantidad", color="tipo", barmode="group")
        fig4a.update_layout(height=360, margin=dict(l=10, r=10, t=10, b=10))
        st.plotly_chart(fig4a, use_container_width=True)

        st.subheader("Gráfico 2 — Hombres vs Mujeres (por provincia)")
        prov_hm = (
            f4.groupby("provincia")[["hombres", "mujeres"]]
            .sum()
            .reset_index()
        )
        prov_hm["__ord__"] = prov_hm["provincia"].apply(lambda x: PROV_ORDER.index(x) if x in PROV_ORDER else 999)
        prov_hm = prov_hm.sort_values("__ord__").drop(columns=["__ord__"])
        prov_hm_long = prov_hm.melt(id_vars=["provincia"], value_vars=["hombres", "mujeres"],
                                    var_name="sexo", value_name="cantidad")

        fig4b = px.bar(prov_hm_long, x="provincia", y="cantidad", color="sexo", barmode="group")
        fig4b.update_layout(height=360, margin=dict(l=10, r=10, t=10, b=10))
        st.plotly_chart(fig4b, use_container_width=True)

        st.subheader("Tabla normalizada")
        tabla4 = (
            f4.groupby(["provincia", "canton"])
            .agg(
                matriculadas=("matriculadas", "sum"),
                egresadas=("egresadas_num", "sum"),
                hombres=("hombres", "sum"),
                mujeres=("mujeres", "sum"),
                cursos=("cursos", lambda s: "\n\n".join(sorted(set(s))))
            )
            .reset_index()
        )
        tabla4 = sort_by_provincia(tabla4)

        st.dataframe(
            title_case_headers(tabla4[["provincia", "canton", "matriculadas", "egresadas", "hombres", "mujeres", "cursos"]]),
            use_container_width=True,
            height=360,
            hide_index=True
        )

        st.markdown("</div>", unsafe_allow_html=True)

    with right:
        st.markdown("<div class='panel'>", unsafe_allow_html=True)
        st.subheader("Mapa satelital (ESRI) — empleabilidad")

        st.markdown("<div class='btnrow'>", unsafe_allow_html=True)
        if st.button("⛶ Ver mapa en pantalla completa", key="t4_full"):
            st.session_state["map_fullscreen_tab4"] = True
            st.rerun()
        st.markdown("<span class='caption'>Popup: matriculadas, egresadas, H/M y cursos brindados.</span>", unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

        render_map_generic(f4.dropna(subset=["lat", "lon"]), height_px=820, popup_mode="empleabilidad")
        st.markdown("</div>", unsafe_allow_html=True)

    st.markdown("<hr/>", unsafe_allow_html=True)

    excel4 = df_to_excel_bytes(tabla4[["provincia", "canton", "matriculadas", "egresadas", "hombres", "mujeres", "cursos"]], sheet_name="Empleabilidad")
    st.download_button(
        "⬇️ Descargar Excel (Empleabilidad)",
        data=excel4,
        file_name="programas_empleabilidad.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_xlsx_t4"
    )
